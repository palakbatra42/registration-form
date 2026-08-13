from ast import Continue
from multiprocessing.reduction import duplicate
from django.shortcuts import render, redirect,get_object_or_404
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from project.Models.Lead import *
from project.Models.LeadForm import *
from project.Controller.Serializers import *
from project.Models.LeadStatusHistory import *
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
import csv
from django.db.models import Count
from django.contrib import messages
from django.shortcuts import redirect, render
from collections import Counter

@login_required(login_url='Login')
def Add_lead(request):
    if request.method == "POST":
        form = LeadForm(request.POST)

        if form.is_valid():
            lead = form.save(commit=False)

            if not lead.assigned_user:
                lead.assigned_user = request.user

            lead.save()

            messages.success(request, "New lead created successfully.")

            return redirect("Dashboard:Record")

        else:
            print(form.errors)

    else:
        form = LeadForm()

    return render(request, "Dashboard/record.html", {"form": form})


@login_required(login_url='Login') 
def Delete_lead(request, id):
    lead = get_object_or_404(Lead, id=id)

    if request.method == "POST":
        lead.delete()
        return redirect("Lead:Record")

    return render(request, "Dashboard/Record.html", {"lead": lead})

@login_required(login_url='Login')
def Update_lead(request, id):
    lead = get_object_or_404(Lead, id=id)

    if request.method == "POST":

        print("POST DATA:", request.POST)
        form = LeadForm(request.POST, instance=lead)

        if form.is_valid():
            print("FORM VALID")
            form.save()
            return redirect("Lead:Record")
        
        else:
            print("ERROR:", form.errors)

    else:
        form = LeadForm(instance=lead)

    return render(request, "Dashboard/Update.html", {
        "form": form,
        "lead": lead,
    })

@login_required(login_url='Login')
def Home(request):
    if request.method == "POST":
        form = LeadForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("Lead:Record")
    else:
        form = LeadForm()
    return render(request, "Register/Home.html", {"form": form})

@login_required(login_url="Login")
def Get_filtered_leads(request): 
    data = Lead.objects.all()

    name = request.GET.get("name")
    phone = request.GET.get("phone")
    email = request.GET.get("email")
    status = request.GET.get("status")
    source = request.GET.get("source")
    assigned_user = request.GET.get("assigned_user")

    # Filters
    if name:
        data = data.filter(name__icontains=name)

    if phone:
        data = data.filter(phone__icontains=phone)

    if email:
        data = data.filter(email__icontains=email)

    if status:
        data = data.filter(status=status)

    if source:
        data = data.filter(source=source)

    if assigned_user:
        data = data.filter(assigned_user=assigned_user)

    # Sorting by complete row using Name
    sort = request.GET.get("sort", "asc")
    
    if sort == "asc":
        data = data.order_by("name")
    elif sort == "desc":
        data = data.order_by("-name")
    return data

   

@login_required(login_url='Login')
def Record(request):
    data = Get_filtered_leads(request)

    print("Logged in user:", request.user.username)
    print("Lead count:", data.count())

    context = Get_Dashboard_Data()

    context.update({
        "data": data,
        "name": request.GET.get("name", ""),
        "status": request.GET.get("status", ""),
        "source": request.GET.get("source", ""),
        "status_choices": Lead.STATUS_CHOICES,
    })

    return render(request, "Dashboard/Record.html", context)

@login_required(login_url='Login')
def Dash(request):
    context = Get_Dashboard_Data()
    return render(request, "Dashboard/Record.html", context)

def Get_Dashboard_Data():
    total_leads = Lead.objects.count()

    open_leads = Lead.objects.filter(
        status__in=[
            "New",
            "Contacted",
            "Qualified",
            "Proposal Sent",
        ]
    ).count()

    converted_leads = Lead.objects.filter(status="Won").count()
    lost_leads = Lead.objects.filter(status="Lost").count()

    if total_leads > 0:
        conversion_rate = round((converted_leads / total_leads) * 100, 2)
    else:
        conversion_rate = 0

    recent_activity = LeadStatusHistory.objects.select_related(
        "lead",
        "changed_by",
    ).order_by("-changed_at")[:10]

    return {
        "total_leads": total_leads,
        "open_leads": open_leads,
        "converted_leads": converted_leads,
        "lost_leads": lost_leads,
        "conversion_rate": conversion_rate,
        "recent_activity": recent_activity,
    }

@login_required(login_url='Login')
def Export_csv(request):
    data = Get_filtered_leads(request)
    columns = request.GET.getlist("columns")

    if not columns:
        columns = ["name", "email", "phone", "source", "status","assigned_user","created_at"]

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="leads.csv"'
    
    writer = csv.writer(response, delimiter=',')
    writer.writerow([col.capitalize() for col in columns])

    for lead in data: 

        row = []

        for col in columns:
            row.append(getattr(lead, col))

        writer.writerow(row)

    return response

@login_required(login_url='Login')
def Export_excel(request):

    data = Get_filtered_leads(request)

    columns = request.GET.getlist("columns")

    if not columns:
        columns = ["name", "email", "phone", "source", "status","assigned_user","created_at"]

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Lead Data"

    sheet.append([col.capitalize() for col in columns])

    for lead in data:

        row = []

        for col in columns:
            row.append(getattr(lead, col))

        sheet.append(row)

    # Summary Sheet
    summary = workbook.create_sheet(title="Summary")

    summary["A1"] = "Total Leads"
    summary["B1"] = data.count()

    summary["A3"] = "Status"
    summary["B3"] = "Count"

    row = 4

    status_counts = data.values("status").annotate(total=Count("id"))

    for item in status_counts:
        summary.cell(row=row, column=1).value = item["status"]
        summary.cell(row=row, column=2).value = item["total"]
        row += 1

    row += 2

    summary.cell(row=row, column=1).value = "Source"
    summary.cell(row=row, column=2).value = "Count"

    row += 1

    source_counts = data.values("source").annotate(total=Count("id"))

    for item in source_counts:
        summary.cell(row=row, column=1).value = item["source"]
        summary.cell(row=row, column=2).value = item["total"]
        row += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="leads.xlsx"'

    workbook.save(response)

    return response

@login_required(login_url='Login')
def Update_status(request, id):
    lead = get_object_or_404(Lead, id=id)

    if request.method == "POST":
        new_status = request.POST.get("status")

        if new_status not in dict(Lead.STATUS_CHOICES):
            messages.error(request, "Invalid status")
            return redirect(request.POST.get("next", "Lead:Record"))

        if lead.status != new_status:
            LeadStatusHistory.objects.create(
                lead=lead,
                old_status=lead.status,
                new_status=new_status,
                changed_by=request.user,
            )
            lead.status = new_status
            lead.save()
            messages.success(request, "Status updated successfully")

    next_url = request.POST.get("next")
    if next_url:
        return redirect(next_url)
    redirect("Lead:Record")

@login_required(login_url='Login')
def Lead_detail(request, id):
    lead = get_object_or_404(Lead, id=id)

    history = LeadStatusHistory.objects.filter(
        lead=lead
    ).order_by("-changed_at")

    return render(request, "Dashboard/Lead_detail.html", {
        "lead": lead,
        "history": history,
    })


def Import_csv(request):
    if request.method == "POST":
        csv_file = request.FILES.get("csv_file")

        if not csv_file:
            messages.error(request, "No file selected")
            return redirect("Lead:Record")

        data = csv_file.read().decode("utf-8").splitlines()
        reader = csv.DictReader(data)

        imported = 0
        duplicate = []
        invalid = []
        names = []

        for row_no,row in enumerate(reader,start=2):
            name = row.get("name", "").strip()
            email = row.get("email", "").strip()
            phone = str(row.get("phone", "")).strip()
            source = row.get("source", "").strip()
            status = row.get("status", "").strip()

            if not name or not email or not phone:
                invalid.append(f"Row {row_no}")
                continue

            if not phone.isdigit():
                invalid.append(f"Row {row_no}")
                continue

            if (name, email, phone) in duplicate:
                duplicate.append((name, email, phone))
                continue
            

            if Lead.objects.filter(email=email,phone = phone).exists():
                duplicate.append((name, email, phone))
                continue

            Lead.objects.create(
                name=name,
                email=email,
                phone=phone,
                source=source,
                status=status if status else "New",
            )

            names.append((name, email, phone))
            imported += 1

        messages.success(request,f"Imported: {imported}, Duplicates: {len(duplicate)}, Invalid: {len(invalid)}")
    return redirect("Lead:Record") 